"""
tests/test_exporters.py — Stage 24 Export Services Test Suite
==============================================================
32 tests covering CSV (stdlib — always runs) and Excel (skipped if openpyxl absent).

Test classes:
  TestCSVExporter      (18 tests) — all six CSV methods, ZIP bundle
  TestExcelExporter    (8  tests) — sheet generation, bytes output  [skipped w/o openpyxl]
  TestExporterEdgeCases(6  tests) — empty runs, zero escalations, unicode payloads
"""

import csv
import io
import sys
import unittest
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from unittest.mock import patch

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_SKIP_XL = unittest.skipUnless(_HAS_OPENPYXL, "openpyxl not installed")


# ── Shared fixture factory ────────────────────────────────────────────────────

def _make_run(
    n_records: int = 3,
    escalate_first: bool = True,
) -> "OrchestratorRun":
    """Create a minimal but structurally complete OrchestratorRun for testing."""
    from core.models import (
        OrchestratorRun, RecordResult, TriageResult, ExecutiveBrief,
        RiskAssessment, EvidenceValidation, RecommendationQuality,
        ExplanationReport, MemoryEntry, GeneratedReport,
        ExplainabilityBlock, RoutingDecision, AgentStatus,
    )

    run = OrchestratorRun(
        run_id="test-run-exporters-001",
        tenant_id="test_tenant",
        model_id="llama-3.1-8b-instant",
        status="completed",
    )

    domains = ["Health", "Education", "Sports", "Politics", "Entertainment"]
    for i in range(n_records):
        score = 9 if (i == 0 and escalate_first) else 4
        routing = (
            RoutingDecision.ESCALATED_TO_AGENTS
            if score >= 7
            else RoutingDecision.BELOW_THRESHOLD
        )
        domain    = domains[i % len(domains)]
        record_id = f"TEST-{i+1:03d}"

        xp = ExplainabilityBlock(
            reasoning="Test reasoning.",
            confidence_score=0.85,
            supporting_evidence=["Evidence A", "Evidence B"],
            risks_of_inaction=["Risk if ignored"],
            expected_outcomes=["Positive outcome"],
        )

        triage = TriageResult(
            record_id=record_id, domain=domain, impact_score=score,
            score_rationale="Test rationale.", primary_risk_flags=["flag1", "flag2"],
            structural_validity="VALID", validity_notes="Complete dataset.",
            urgency="CRITICAL" if score >= 9 else "LOW",
            stakeholder_tier="Clinical Staff" if domain == "Health" else "General",
            data_freshness="REAL-TIME",
            triage_summary=f"Test summary for {domain}.",
            explainability=xp, agent_status=AgentStatus.SUCCESS,
            execution_time_ms=1200,
        )

        rec = RecordResult(
            record_id=record_id, domain=domain,
            processing_order=i + 1,
            routing_decision=routing,
            triage=triage,
            total_execution_ms=5000,
        )

        if routing == RoutingDecision.ESCALATED_TO_AGENTS:
            rec.executive = ExecutiveBrief(
                record_id=record_id, domain=domain,
                executive_brief="Critical situation requiring immediate executive attention.",
                recommended_action="Deploy emergency response team within 2 hours.",
                action_link="https://who.int/emergencies",
                escalation_tier="C-SUITE",
                response_deadline="Within 2 hours",
                key_metrics=["Score: 9/10", "Urgency: CRITICAL"],
                stakeholders=["CMO", "CEO"],
                agent_status=AgentStatus.SUCCESS,
            )
            rec.risk = RiskAssessment(
                record_id=record_id, domain=domain,
                overall_risk_level="CRITICAL", risk_score=9.2,
                risk_categories={
                    "operational": "CRITICAL", "financial": "HIGH",
                    "reputational": "HIGH", "regulatory": "CRITICAL",
                    "strategic": "MODERATE",
                },
                mitigation_steps=["Step 1: Isolate", "Step 2: Escalate", "Step 3: Resolve"],
                residual_risks=["Residual risk after mitigation"],
                agent_status=AgentStatus.SUCCESS,
            )
            rec.evidence = EvidenceValidation(
                record_id=record_id, domain=domain,
                validation_status="VERIFIED", data_quality_score=0.92,
                verified_claims=["Claim A verified", "Claim B verified"],
                agent_status=AgentStatus.SUCCESS,
            )
            rec.report = GeneratedReport(
                record_id=record_id, domain=domain,
                executive_summary="High-impact situation detected requiring immediate action.",
                board_summary="Board must convene emergency session.",
                operational_summary="Operations team to mobilise within the hour.",
                key_findings=["Finding 1", "Finding 2", "Finding 3"],
                recommendations=["Recommendation A", "Recommendation B"],
                next_steps=["Immediate: Alert CMO", "Short-term: Draft response plan"],
                agent_status=AgentStatus.SUCCESS,
            )

        run.records.append(rec)

    run.completed_at = "2026-01-01T12:00:00+00:00"
    run.build_summary()
    return run


# ══════════════════════════════════════════════════════════════════════════════
# CSV EXPORTER
# ══════════════════════════════════════════════════════════════════════════════
class TestCSVExporter(unittest.TestCase):

    def setUp(self):
        from services.csv_exporter import CSVExporter
        self.exporter = CSVExporter()
        self.run      = _make_run(n_records=3, escalate_first=True)

    # ── domain_scores_csv ─────────────────────────────────────────────────────

    def test_domain_scores_returns_string(self):
        result = self.exporter.domain_scores_csv(self.run)
        self.assertIsInstance(result, str)

    def test_domain_scores_has_header(self):
        result = self.exporter.domain_scores_csv(self.run)
        reader = csv.DictReader(io.StringIO(result))
        self.assertIn("record_id", reader.fieldnames)
        self.assertIn("impact_score", reader.fieldnames)
        self.assertIn("urgency", reader.fieldnames)

    def test_domain_scores_row_count_matches_records(self):
        result = self.exporter.domain_scores_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(len(rows), len(self.run.records))

    def test_domain_scores_escalated_record_has_correct_score(self):
        result = self.exporter.domain_scores_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        first  = rows[0]
        self.assertEqual(first["impact_score"], "9")
        self.assertEqual(first["urgency"], "CRITICAL")

    # ── risk_matrix_csv ───────────────────────────────────────────────────────

    def test_risk_matrix_returns_string(self):
        result = self.exporter.risk_matrix_csv(self.run)
        self.assertIsInstance(result, str)

    def test_risk_matrix_only_includes_records_with_risk(self):
        result = self.exporter.risk_matrix_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        # Only the escalated record has a risk assessment in our fixture
        self.assertEqual(len(rows), 1)

    def test_risk_matrix_has_category_columns(self):
        result = self.exporter.risk_matrix_csv(self.run)
        reader = csv.DictReader(io.StringIO(result))
        for cat in ["operational", "financial", "reputational", "regulatory", "strategic"]:
            self.assertIn(cat, reader.fieldnames)

    def test_risk_matrix_overall_risk_correct(self):
        result = self.exporter.risk_matrix_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(rows[0]["overall_risk"], "CRITICAL")

    # ── escalations_csv ───────────────────────────────────────────────────────

    def test_escalations_returns_string(self):
        result = self.exporter.escalations_csv(self.run)
        self.assertIsInstance(result, str)

    def test_escalations_only_includes_escalated(self):
        result = self.exporter.escalations_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(len(rows), 1)      # only one escalated in fixture

    def test_escalations_has_action_fields(self):
        result = self.exporter.escalations_csv(self.run)
        reader = csv.DictReader(io.StringIO(result))
        self.assertIn("recommended_action", reader.fieldnames)
        self.assertIn("action_link", reader.fieldnames)
        self.assertIn("escalation_tier", reader.fieldnames)

    # ── recommendations_csv ───────────────────────────────────────────────────

    def test_recommendations_returns_string(self):
        result = self.exporter.recommendations_csv(self.run)
        self.assertIsInstance(result, str)

    def test_recommendations_numbering_starts_at_1(self):
        result = self.exporter.recommendations_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        if rows:
            self.assertEqual(rows[0]["priority"], "1")

    def test_recommendations_contains_content(self):
        result = self.exporter.recommendations_csv(self.run)
        self.assertIn("Recommendation A", result)

    # ── audit_trail_csv ───────────────────────────────────────────────────────

    def test_audit_trail_returns_string(self):
        result = self.exporter.audit_trail_csv(self.run)
        self.assertIsInstance(result, str)

    def test_audit_trail_has_header_columns(self):
        result = self.exporter.audit_trail_csv(self.run)
        reader = csv.DictReader(io.StringIO(result))
        for col in ["timestamp", "event_type", "severity"]:
            self.assertIn(col, reader.fieldnames)

    # ── summary_csv ───────────────────────────────────────────────────────────

    def test_summary_csv_returns_exactly_one_data_row(self):
        result = self.exporter.summary_csv(self.run)
        rows   = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(len(rows), 1)

    # ── ZIP bundle ────────────────────────────────────────────────────────────

    def test_zip_bundle_returns_bytes(self):
        result = self.exporter.export_all_as_zip(self.run)
        self.assertIsInstance(result, bytes)

    def test_zip_bundle_is_valid_zip(self):
        result = self.exporter.export_all_as_zip(self.run)
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(result)))

    def test_zip_bundle_contains_all_six_files(self):
        result = self.exporter.export_all_as_zip(self.run)
        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            names = zf.namelist()
        self.assertEqual(len(names), 6)
        joined = " ".join(names)
        for keyword in ["domain_scores", "risk_matrix", "escalations",
                        "recommendations", "audit_trail", "summary"]:
            self.assertIn(keyword, joined)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════
@_SKIP_XL
class TestExcelExporter(unittest.TestCase):

    def setUp(self):
        from services.excel_exporter import ExcelExporter
        self.exporter = ExcelExporter()
        self.run      = _make_run(n_records=3, escalate_first=True)

    def test_export_returns_bytes(self):
        result = self.exporter.export_run(self.run)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_export_produces_valid_xlsx(self):
        result = self.exporter.export_run(self.run)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        self.assertIsNotNone(wb)

    def test_workbook_has_all_required_sheets(self):
        result = self.exporter.export_run(self.run)
        wb     = openpyxl.load_workbook(io.BytesIO(result))
        expected = {"Summary", "Domain Scores", "Risk Matrix",
                    "Escalations", "Recommendations", "Audit Trail"}
        self.assertEqual(set(wb.sheetnames), expected)

    def test_domain_scores_sheet_has_data_rows(self):
        result = self.exporter.export_run(self.run)
        wb     = openpyxl.load_workbook(io.BytesIO(result))
        ws     = wb["Domain Scores"]
        # Row 1 = header; rows 2+ = data
        self.assertGreaterEqual(ws.max_row, 2)

    def test_summary_sheet_has_kpi_labels(self):
        result = self.exporter.export_run(self.run)
        wb     = openpyxl.load_workbook(io.BytesIO(result))
        ws     = wb["Summary"]
        cell_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        self.assertIn("Run ID", cell_values)

    def test_escalations_sheet_only_contains_escalated(self):
        result = self.exporter.export_run(self.run)
        wb     = openpyxl.load_workbook(io.BytesIO(result))
        ws     = wb["Escalations"]
        # Header + 1 escalated record
        self.assertEqual(ws.max_row, 2)

    def test_save_to_file(self):
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        try:
            saved = self.exporter.export_run_to_file(self.run, path=tmp_path)
            self.assertTrue(saved.exists())
            self.assertGreater(saved.stat().st_size, 0)
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_no_openpyxl_raises_runtime_error(self):
        """Verify helpful error when openpyxl is absent."""
        import unittest.mock as mock
        with mock.patch("services.excel_exporter._HAS_OPENPYXL", False):
            from services.excel_exporter import ExcelExporter as _EX
            with self.assertRaises(RuntimeError):
                _EX()


# ══════════════════════════════════════════════════════════════════════════════
# EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════
class TestExporterEdgeCases(unittest.TestCase):

    def setUp(self):
        from services.csv_exporter import CSVExporter
        self.exporter = CSVExporter()

    def test_empty_run_domain_scores_has_only_header(self):
        run   = _make_run(n_records=0)
        result = self.exporter.domain_scores_csv(run)
        rows  = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(rows, [])

    def test_run_with_no_escalations_gives_empty_escalations_csv(self):
        run   = _make_run(n_records=2, escalate_first=False)
        result = self.exporter.escalations_csv(run)
        rows  = list(csv.DictReader(io.StringIO(result)))
        self.assertEqual(rows, [])

    def test_unicode_payload_survives_csv_round_trip(self):
        from core.models import OrchestratorRun, RecordResult, TriageResult, RoutingDecision, AgentStatus
        run = OrchestratorRun(run_id="unicode-test", tenant_id="t", model_id="m", status="completed")
        t = TriageResult(
            record_id="U-001", domain="政治", impact_score=3,
            score_rationale="テスト", primary_risk_flags=["flag"],
            structural_validity="VALID", validity_notes="",
            urgency="LOW", stakeholder_tier="", data_freshness="RECENT",
            triage_summary="Résumé en français avec des accents: café, naïve, über.",
            agent_status=AgentStatus.SUCCESS,
        )
        run.records.append(RecordResult(
            record_id="U-001", domain="政治", processing_order=1,
            routing_decision=RoutingDecision.BELOW_THRESHOLD, triage=t,
        ))
        run.build_summary()
        result = self.exporter.domain_scores_csv(run)
        self.assertIn("政治", result)

    def test_zip_empty_run_still_produces_valid_zip(self):
        run    = _make_run(n_records=0)
        result = self.exporter.export_all_as_zip(run)
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(result)))

    def test_summary_csv_includes_run_id(self):
        run    = _make_run(n_records=1)
        result = self.exporter.summary_csv(run)
        self.assertIn(run.run_id, result)

    def test_audit_filter_by_severity(self):
        run    = _make_run(n_records=1)
        result = self.exporter.audit_trail_csv(run, filter_severity="ERROR")
        rows   = list(csv.DictReader(io.StringIO(result)))
        for row in rows:
            self.assertEqual(row["severity"], "ERROR")


if __name__ == "__main__":
    unittest.main(verbosity=2)
