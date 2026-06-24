"""
services/excel_exporter.py — Enterprise Excel Report Generator
===============================================================
Produces a multi-sheet XLSX workbook from an OrchestratorRun.

Sheets:
  1. Summary      — KPI overview and run metadata
  2. Domain Scores — Impact scores, urgency, routing decision per record
  3. Risk Matrix   — Risk levels across all risk categories per domain
  4. Escalations   — Full detail for every escalated record
  5. Recommendations — Prioritised action list
  6. Audit Trail   — Last 100 audit events

Visual style:
  • Navy header rows with white bold text  (#0F172A)
  • Alternating light row shading
  • Colour-coded urgency cells
  • Column widths auto-fitted to content
  • Frozen top row on every sheet

Dependency: openpyxl >= 3.1.0  (pip install openpyxl)
Falls back gracefully to a not-available message if openpyxl is absent.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from config import config
from core.logger import get_logger
from core.models import OrchestratorRun, RoutingDecision

log = get_logger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    log.warning("openpyxl not installed — Excel export unavailable. pip install openpyxl")

# ── Colour palette ─────────────────────────────────────────────────────────────
_NAVY   = "0F172A"
_INDIGO = "6366F1"
_WHITE  = "FFFFFF"
_LIGHT  = "F8FAFC"
_ALT    = "E2E8F0"

_URGENCY_FILL = {
    "CRITICAL": "450A0A",
    "HIGH":     "431407",
    "MEDIUM":   "172554",
    "LOW":      "052E16",
}
_URGENCY_FONT = {
    "CRITICAL": "FCA5A5",
    "HIGH":     "FCD34D",
    "MEDIUM":   "93C5FD",
    "LOW":      "86EFAC",
}


class ExcelExporter:
    """
    Stateless Excel builder.
    Call export_run() to produce bytes for st.download_button or file saving.
    """

    def __init__(self) -> None:
        if not _HAS_OPENPYXL:
            raise RuntimeError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

    # ── Public ────────────────────────────────────────────────────────────────

    def export_run(self, run: OrchestratorRun) -> bytes:
        """Return a complete XLSX workbook as raw bytes."""
        wb = Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        self._sheet_summary(wb, run)
        self._sheet_domain_scores(wb, run)
        self._sheet_risk_matrix(wb, run)
        self._sheet_escalations(wb, run)
        self._sheet_recommendations(wb, run)
        self._sheet_audit(wb, run)

        buf = io.BytesIO()
        wb.save(buf)
        pdf_bytes = buf.getvalue()
        buf.close()

        log.info(
            "Excel workbook generated",
            extra={"run_id": run.run_id, "sheets": len(wb.sheetnames),
                   "bytes": len(pdf_bytes)},
        )
        return pdf_bytes

    def export_run_to_file(
        self, run: OrchestratorRun, path: Optional[Path] = None
    ) -> Path:
        if path is None:
            config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
            ts   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            path = config.REPORTS_DIR / f"report_{run.run_id[:8]}_{ts}.xlsx"
        path.write_bytes(self.export_run(run))
        return path

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _sheet_summary(self, wb: "Workbook", run: OrchestratorRun) -> None:
        ws = wb.create_sheet("Summary")
        summary = run.summary or {}

        self._header_row(ws, 1, ["Enterprise Decision Intelligence Report"])
        ws.merge_cells("A1:F1")
        ws["A1"].font = Font(bold=True, color=_WHITE, size=14)

        # Metadata block
        meta = [
            ["Run ID",    run.run_id],
            ["Tenant",    run.tenant_id],
            ["Model",     run.model_id],
            ["Status",    run.status.upper()],
            ["Started",   run.run_timestamp],
            ["Completed", run.completed_at or "—"],
        ]
        for i, (label, value) in enumerate(meta, start=3):
            ws.cell(i, 1, label).font = Font(bold=True)
            ws.cell(i, 2, str(value))

        # KPI block
        ws.cell(10, 1, "KPI Summary").font = Font(bold=True, color=_NAVY, size=12)
        kpi_headers = ["Records", "Escalated", "Below Threshold", "Errors", "Tokens Used"]
        kpi_values  = [
            summary.get("total_records_processed", 0),
            summary.get("escalated_records", 0),
            summary.get("below_threshold_records", 0),
            summary.get("errors_encountered", 0),
            summary.get("total_tokens_used", 0),
        ]
        self._header_row(ws, 11, kpi_headers)
        for j, v in enumerate(kpi_values, start=1):
            c = ws.cell(12, j, v)
            c.alignment = Alignment(horizontal="center")

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_domain_scores(self, wb: "Workbook", run: OrchestratorRun) -> None:
        ws = wb.create_sheet("Domain Scores")
        headers = ["Record ID", "Domain", "Impact Score", "Urgency",
                   "Routing Decision", "Risk Level", "Confidence", "Execution (ms)"]
        self._header_row(ws, 1, headers)

        for i, rec in enumerate(run.records, start=2):
            t = rec.triage
            bg, fg = "FFFFFF", "000000"
            urgency = t.urgency if t else "UNKNOWN"
            if urgency in _URGENCY_FILL:
                bg, fg = _URGENCY_FILL[urgency], _URGENCY_FONT[urgency]

            row_data = [
                rec.record_id,
                rec.domain,
                t.impact_score if t else 0,
                urgency,
                rec.routing_decision,
                rec.risk.overall_risk_level if rec.risk else "N/A",
                f"{t.explainability.confidence_score:.0%}" if t and t.explainability else "N/A",
                rec.total_execution_ms,
            ]
            for j, v in enumerate(row_data, start=1):
                c = ws.cell(i, j, v)
                if j == 4:   # urgency cell
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.font = Font(bold=True, color=fg)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", fgColor=_ALT)

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_risk_matrix(self, wb: "Workbook", run: OrchestratorRun) -> None:
        ws = wb.create_sheet("Risk Matrix")
        cats = ["operational", "financial", "reputational", "regulatory", "strategic"]
        headers = ["Record ID", "Domain", "Overall Risk", "Risk Score"] + [c.title() for c in cats]
        self._header_row(ws, 1, headers)

        for i, rec in enumerate(run.records, start=2):
            if not rec.risk:
                continue
            row = [
                rec.record_id, rec.domain,
                rec.risk.overall_risk_level,
                rec.risk.risk_score,
            ] + [rec.risk.risk_categories.get(c, "N/A") for c in cats]
            for j, v in enumerate(row, start=1):
                ws.cell(i, j, v)
            # Colour overall risk cell
            level = rec.risk.overall_risk_level
            level_fill = {"CRITICAL": "450A0A", "HIGH": "431407",
                          "MODERATE": "172554", "LOW": "052E16"}.get(level, "FFFFFF")
            level_font = {"CRITICAL": "FCA5A5", "HIGH": "FCD34D",
                          "MODERATE": "93C5FD", "LOW": "86EFAC"}.get(level, "000000")
            ws.cell(i, 3).fill = PatternFill("solid", fgColor=level_fill)
            ws.cell(i, 3).font = Font(bold=True, color=level_font)

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_escalations(self, wb: "Workbook", run: OrchestratorRun) -> None:
        ws = wb.create_sheet("Escalations")
        headers = ["Record ID", "Domain", "Impact Score", "Executive Brief",
                   "Recommended Action", "Action Link", "Escalation Tier",
                   "Deadline", "Validation Status"]
        self._header_row(ws, 1, headers)

        row_num = 2
        for rec in run.records:
            if rec.routing_decision != RoutingDecision.ESCALATED_TO_AGENTS:
                continue
            e = rec.executive
            v = rec.evidence
            row = [
                rec.record_id, rec.domain,
                rec.triage.impact_score if rec.triage else 0,
                (e.executive_brief[:200] if e and e.executive_brief else "N/A"),
                (e.recommended_action    if e else "N/A"),
                (e.action_link           if e else "N/A"),
                (e.escalation_tier       if e else "N/A"),
                (e.response_deadline     if e else "N/A"),
                (v.validation_status     if v else "N/A"),
            ]
            for j, v_ in enumerate(row, start=1):
                c = ws.cell(row_num, j, str(v_))
                c.alignment = Alignment(wrap_text=True)
            row_num += 1

        self._auto_width(ws)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    def _sheet_recommendations(self, wb: "Workbook", run: OrchestratorRun) -> None:
        ws = wb.create_sheet("Recommendations")
        self._header_row(ws, 1, ["#", "Record", "Domain", "Recommendation", "Next Step"])
        counter, row_num = 1, 2
        for rec in run.records:
            if not rec.report:
                continue
            for r in rec.report.recommendations:
                ws.cell(row_num, 1, counter)
                ws.cell(row_num, 2, rec.record_id)
                ws.cell(row_num, 3, rec.domain)
                ws.cell(row_num, 4, r).alignment = Alignment(wrap_text=True)
                ws.cell(row_num, 5, rec.report.next_steps[0] if rec.report.next_steps else "")
                counter += 1
                row_num += 1
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_audit(self, wb: "Workbook", run: OrchestratorRun) -> None:
        from core.logger import AuditLogger
        ws = wb.create_sheet("Audit Trail")
        self._header_row(ws, 1, ["Timestamp", "Event Type", "Severity",
                                  "Agent", "Record", "Detail"])
        entries = AuditLogger.read_all()
        for i, e in enumerate(entries[-200:], start=2):
            data  = e.get("data", {})
            detail = str(data)[:120] if isinstance(data, dict) else str(data)[:120]
            row = [
                e.get("timestamp", "")[:19].replace("T", " "),
                e.get("event_type", ""),
                e.get("severity", ""),
                e.get("agent_name", ""),
                e.get("record_id", ""),
                detail,
            ]
            for j, v in enumerate(row, start=1):
                c = ws.cell(i, j, str(v))
                sev = e.get("severity", "")
                if sev == "ERROR":
                    c.font = Font(color="EF4444")
                elif sev == "WARN":
                    c.font = Font(color="F59E0B")
        self._auto_width(ws)
        ws.freeze_panes = "A2"

    # ── Helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _header_row(ws, row: int, cols: list) -> None:
        for j, label in enumerate(cols, start=1):
            c = ws.cell(row, j, label)
            c.fill = PatternFill("solid", fgColor=_NAVY)
            c.font = Font(bold=True, color=_WHITE)
            c.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _auto_width(ws, max_width: int = 60) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
